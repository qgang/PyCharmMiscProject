import pandas as pd
import sys
from openpyxl.styles import Alignment


def read_excel_file(file_path):
    try:
        # 关键修正：读取.xls文件必须指定engine='xlrd'
        df = pd.read_excel(file_path, engine='xlrd')
        # 检查读取结果是否为空
        if df.empty:
            print("错误: 读取的Excel文件为空")
            return None
        return df
    except FileNotFoundError:
        print("错误: 文件未找到，请检查文件路径是否正确。")
        return None
    except ImportError:
        print("错误: 缺少xlrd库，请执行 pip install xlrd>=2.0.1 -i https://pypi.tuna.tsinghua.edu.cn/simple")
        return None
    except Exception as e:
        print(f"错误: 读取文件时发生未知错误: {e}")
        return None


def handle_score(score):
    # 鲁棒性检查：确保行数足够（至少3行：表头1+表头2+数据）
    if len(score) < 3:
        print("错误: 分数表行数不足，无法处理表头")
        return None

    try:
        # 处理表头：删除第一行，第二行作为新表头
        score.iloc[1, :4] = score.iloc[0, :4].values  # 加.values避免索引对齐问题
        new_score = score.iloc[2:].copy()  # 用iloc更安全，加copy避免SettingWithCopyWarning
        new_score.columns = score.iloc[1].values  # 用.values确保列名是列表而非Series
        new_score = new_score.reset_index(drop=True)  # 重置索引

        # 删除包含"答案"的列和"学号"列
        target_str = '答案'
        columns_to_drop = [col for col in new_score.columns if isinstance(col, str) and target_str in col]
        new_score = new_score.drop(columns=columns_to_drop, errors='ignore')  # errors='ignore'避免列不存在报错
        new_score = new_score.drop(columns='学号', errors='ignore')

        # 重命名列名：截断"（"后的内容
        rename_dict = {}
        for col in new_score.columns:
            if isinstance(col, str):  # 确保列名是字符串
                idx = col.find('（')
                rename_dict[col] = col[:idx] if idx != -1 else col
            else:
                rename_dict[col] = str(col)  # 非字符串列名转为字符串

        new_score = new_score.rename(columns=rename_dict)

        # 排序：检查"考号"列是否存在
        if '考号' not in new_score.columns:
            print("错误: 数据中未找到'考号'列，无法排序")
            return None

        # 排序时忽略NaN值
        new_score = new_score.sort_values(by='考号', na_position='last')
        return new_score
    except Exception as e:
        print(f"处理分数表时出错: {e}")
        return None


def handle_point(score_sheet):
    if score_sheet is None or score_sheet.empty:
        print("错误: 分数表为空，无法生成考点划分表")
        return None

    try:
        # 修正：空列名改为有意义的名称，避免后续问题
        columns = ['题型/题号', '全卷', '1卷', '2卷', '听力', '语法填空', '选词填空', '完型填空', '阅读', '六选四',
                   '概要', '翻译', '作文']
        # 获取需要填充到"题型/题号"列的内容（score_sheet第3列之后的列名）
        question_cols = score_sheet.columns[3:].tolist()  # 转为列表，避免可迭代对象异常
        if not question_cols:
            print("错误: 分数表有效列数不足，无法生成考点划分表")
            return None

        # 初始化考点表：行数匹配question_cols的长度
        point_sheet = pd.DataFrame(
            index=range(len(question_cols)),
            columns=columns
        )
        # 填充"题型/题号"列
        point_sheet['题型/题号'] = question_cols

        # 修正：变量名不重复（sections → section）
        sections = {
            '全卷': (0, 1),
            '1卷': (1, 2),
            '2卷': (2, 3)
        }
        # 给全卷/1卷/2卷赋值：检查行索引是否有效
        for section, (row_idx, col_idx) in sections.items():
            if row_idx < len(point_sheet):  # 避免行索引越界
                point_sheet.iloc[row_idx, col_idx] = 1

        # 处理听力列：只给存在的行赋值（避免创建NaN行）
        hearing_row_start = 3
        hearing_row_end = min(22, len(point_sheet) - 1)  # 不超过实际行数
        if hearing_row_start <= hearing_row_end:
            point_sheet.loc[hearing_row_start:hearing_row_end, '听力'] = 1

        # 填充NaN为空白（避免浮点数NaN引发迭代报错）
        point_sheet = point_sheet.fillna('')
        return point_sheet
    except Exception as e:
        print(f"生成考点划分表时出错: {e}")
        return None


def save_to_excel(point_sheet, output_file):
    if point_sheet is None or point_sheet.empty:
        print("错误: 考点划分表为空，无法保存")
        return False

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            point_sheet.to_excel(writer, sheet_name="考点划分表", index=False)

            # 获取工作簿和工作表对象
            wb = writer.book
            ws = writer.sheets["考点划分表"]

            # 冻结首行
            ws.freeze_panes = "A2"

            # 设置单元格居中对齐：遍历所有有数据的单元格
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        print(f"考点划分表已成功保存到: {output_file}")
        return True
    except Exception as e:
        print(f"保存Excel失败: {e}")
        return False


if __name__ == "__main__":
    input_file = './score.xls'
    output_file = './point.xlsx'

    try:
        print("读取分数表......")
        score = read_excel_file(input_file)
        if score is None:
            print("分数表读取失败")
            sys.exit(1)

        print("处理分数表......")
        score_sheet = handle_score(score)
        if score_sheet is None:
            print("分数表处理失败")
            sys.exit(1)

        print("生成考点划分表......")
        point_sheet = handle_point(score_sheet)
        if point_sheet is None:
            print("考点划分表生成失败")
            sys.exit(1)

        # 保存文件
        if not save_to_excel(point_sheet, output_file):
            sys.exit(1)

        print("考点划分初表生成结束！")
    except Exception as e:
        print(f"程序执行出错: {e}")
        sys.exit(1)