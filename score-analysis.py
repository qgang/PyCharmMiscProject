import sys
from zoneinfo import ZoneInfo

import pandas as pd
import numpy as np
from openpyxl.comments import Comment
from datetime import datetime, timezone

from openpyxl.styles import Alignment


def read_excel_file(file_path):
    try:
        # 读取 Excel 文件
        df = pd.read_excel(file_path)
        return df
    except FileNotFoundError as e:
        print("错误: 文件未找到，请检查文件路径是否正确。", e)
    except Exception as e:
        print(f"错误: 发生了一个未知错误: {e}")
    return None

def handle_score(score):
    # 处理表头，删除第一行，且将第二行数据作为新表头
    score.iloc[1, :4] = score.iloc[0, :4]
    new_score = score[2:]
    new_score.columns = score.iloc[1]
    new_score = new_score.reset_index(drop=True)  # 重置索引

    # 处理列
    target_str = '答案'
    columns_to_drop = [col for col in new_score.columns if target_str in col]
    new_score = new_score.drop(columns=columns_to_drop)
    new_score = new_score.drop(columns='学号')

    # 重命名header
    rename_dict = {}
    for col in new_score.columns:
        idx = col.find('（')
        rename_dict[col] = col[:idx] if idx != -1 else col

    new_score = new_score.rename(columns=rename_dict)

    # 排序
    return new_score.sort_values(by='考号')

def handle_point(point):
    new_columns = point.columns.tolist()
    new_columns[0] = ' '
    point.columns = new_columns
    return point

def handle_point_score(point_sheet, score_sheet):
    columns = score_sheet.columns.tolist()[0:3] + point_sheet.columns.tolist()[1::]
    point_score_sheet = pd.DataFrame(columns=columns, index=score_sheet.index)

    # 复制学生信息列
    point_score_sheet.iloc[:, :6] = score_sheet.iloc[:, :6]

    for col in point_sheet.columns[4:]:
        point = point_sheet.loc[point_sheet[col] == 1, point_sheet.columns[0]].tolist()
        point_score_sheet[col] = score_sheet[point].sum(axis=1)

    return point_score_sheet

def handle_student_score(point_score_sheet):
    # 复制学生信息列(前3列)
    student_score = point_score_sheet.iloc[:, :3].copy()

    # 计算每列的平均值并格式化分数为"实际得分/平均分"
    for col in point_score_sheet.columns[3:]:
        mean_val = point_score_sheet[col].mean()
        student_score[col] = point_score_sheet[col].apply(lambda x: f"{x}/{mean_val:.1f}")
        #student_score[col] = [f"{val}/{'%.1f'%mean}" for val in point_score_sheet[col]]

    # 计算扩展后数据的总行数
    total_rows = len(student_score) * 4
    wide = 10

    # 创建结果DataFrame
    student_score_sheet = pd.DataFrame(
        columns=np.arange(wide),  # 0-12共13列
        index=np.arange(total_rows)
    )

    # 处理每个学生的数据
    for i, (_, row) in enumerate(student_score.iterrows()):
        base_idx = i * 4
        columns = student_score.columns

        # 计算宽列和长列的分割点
        split_point = min(wide-1, len(columns) - 3)  # 确保不越界

        # 第一行：班级 + 前部分考点
        student_score_sheet.iloc[base_idx, 0] = row.iloc[2]  # 班级
        student_score_sheet.iloc[base_idx, 1:split_point + 1] = columns[3:3 + split_point]

        # 第二行：姓名 + 前部分分数
        student_score_sheet.iloc[base_idx + 1, 0] = row.iloc[1]  # 姓名
        student_score_sheet.iloc[base_idx + 1, 1:split_point + 1] = row[3:3 + split_point]

        # 如果有剩余考点，添加到第三、四行
        remaining = len(columns) - 3 - split_point
        if remaining > 0:
            # 第三行：剩余考点
            student_score_sheet.iloc[base_idx + 2, 1:remaining + 1] = columns[3 + split_point:]

            # 第四行：剩余分数
            student_score_sheet.iloc[base_idx + 3, 1:remaining + 1] = row[3 + split_point:]

    return student_score_sheet


def handle_class_score(point_score_sheet):
    # 基本数据准备
    total_students = len(point_score_sheet)
    classes = point_score_sheet['行政班级'].unique()
    point_columns = point_score_sheet.columns[3:]  # 假设前3列是学生信息，后面是题目得分

    # 预先计算每个班级的学生数量
    class_student_counts = point_score_sheet['行政班级'].value_counts()

    # 初始化结果数据列表
    class_score_data = []

    # 遍历每个题目
    for point_col in point_columns:
        # 获取该题目的所有不同得分，并按降序排列
        unique_scores = sorted(point_score_sheet[point_col].dropna().unique(), reverse=True)

        # 遍历每个得分
        for score in unique_scores:
            # 筛选出该题目得分为score的所有学生
            score_students = point_score_sheet[point_score_sheet[point_col] == score]

            # 初始化行数据：题号和得分
            row_data = [point_col, score]

            # 计算每个班级的比例
            for class_name in classes:
                # 筛选该班级得分为score的学生
                class_score_students = score_students[score_students['行政班级'] == class_name]

                # 计算比例并格式化
                if len(class_score_students) == 0:
                    row_data.append(' ')  # 没有学生得此分数
                else:
                    # 计算百分比并四舍五入为整数
                    percentage = len(class_score_students) / class_student_counts[class_name]
                    row_data.append(percentage)

            # 计算总比例
            total_percentage = len(score_students) / total_students
            row_data.append(total_percentage)

            # 添加到结果数据
            class_score_data.append(row_data)

        # 计算各班平均分
        row_data = [point_col, '均分']
        class_avg = point_score_sheet.groupby('行政班级')[point_col].mean()
        for clss in classes:
            row_data.append(round(class_avg[clss], 1))

        # 计算总平均分
        total_avg = point_score_sheet[point_col].mean().round(1)
        row_data.append(total_avg)
        class_score_data.append(row_data)

    # 创建结果DataFrame
    result_columns = ['题号', '得分']
    result_columns.extend(classes)
    result_columns.append('合计')

    class_score_sheet = pd.DataFrame(class_score_data, columns=result_columns)

    return class_score_sheet

if __name__ == "__main__":
    print()
    print(datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S") + " 读取分数表......")
    score = read_excel_file('./score.xls')
    if score is None:
        print("分数表读取失败")
        sys.exit(1)
    score_sheet = handle_score(score)


    print(datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S") + " 读取考点划分表......")
    point = read_excel_file('./point.xlsx')
    if point is None:
        print("考点划分读取失败")
        sys.exit(1)
    point_sheet = handle_point(point)

    print(datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S") + " 生成考点分数表......")
    point_score_sheet = handle_point_score(point_sheet, score_sheet)

    print(datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S") + " 生成学生得分清单表......")
    student_score_sheet = handle_student_score(point_score_sheet)

    print(datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S") + " 生成班级得分率表......")
    class_score_sheet = handle_class_score(point_score_sheet)

    output_file = './score_analysis.xlsx'
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            score_sheet.to_excel(writer, sheet_name='分数表', index=False)
            point_sheet.to_excel(writer, sheet_name='考点划分表', index=False)
            point_score_sheet.to_excel(writer, sheet_name="考点分数表", index=False)
            student_score_sheet.to_excel(writer, sheet_name="学生得分清单表", index=False, header=False)
            class_score_sheet.to_excel(writer, sheet_name="班级得分率表", index=False)

            # 冻结工作表的首行
            for sheet_name in ['分数表', '考点划分表', '考点分数表', '班级得分率表']:
                ws = writer.sheets[sheet_name]
                ws.freeze_panes = "A2"

            # 学生得分清单表居中显示
            ws = writer.sheets['学生得分清单表']
            # 获取最大行数和列数
            max_row = ws.max_row
            max_col = ws.max_column
            center_alignment = Alignment(horizontal='center', vertical='center')
            # 遍历所有单元格并应用居中对齐
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.alignment = center_alignment

            # 处理班级得分率表的批注和格式
            ws = writer.sheets["班级得分率表"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1,
                                    min_col=2, max_col=ws.max_column):
                row_size = len(row)
                if row[0].value == '均分':
                    continue

                score = row[0].value
                for cell in row[1:-1]:
                    if not isinstance(cell.value, (int, float)):
                        continue

                    # 获取班级和考点名称
                    class_name = ws.cell(row=1, column=cell.column).value
                    point_name = ws.cell(row=cell.row, column=1).value

                    students_class = point_score_sheet[(point_score_sheet['行政班级'] == class_name)]
                    students_score = students_class[(students_class[point_name] == score)]
                    cell.comment = Comment(", ".join(students_score['姓名'].values), 'sys')
                    cell.number_format = '0%'
                # 设置最后一列格式
                row[row_size - 1].number_format = '0%'

    except Exception as e:
        print("错误: 发生了一个未知错误:", e)

    print(datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S") + " 成绩分析结束！")
